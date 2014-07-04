from django.conf import settings
from openpyxl import Workbook

class libExcel:
    def generateWeekReport(self, data, outputfile):
        book        = Workbook()
        sheet       = book.get_active_sheet()
        n           = 1
        i           = 1
        sheet.title = "Reporte de Actividades"
        column       =sheet.cell(row=n, column=i)
        column.value = "PROYECTO"
        i  = i + 1
        column2      =sheet.cell(row=n, column=i)
        column2.value="TAREA"
        i  = i + 1
        column3      =sheet.cell(row=n, column=i)
        column3.value="LUNES"
        i  = i + 1
        column4      =sheet.cell(row=n, column=i)
        column4.value="MARTES"
        i  = i + 1
        column5      =sheet.cell(row=n, column=i)
        column5.value="MIERCOLES"
        i  = i + 1
        column6      =sheet.cell(row=n, column=i)
        column6.value="JUEVES"
        i  = i + 1
        column7      =sheet.cell(row=n, column=i)
        column7.value="VIERNES"
        i  = i + 1
        column8      =sheet.cell(row=n, column=i)
        column8.value="SABADO"
        i  = i + 1
        
        for project in data:
           
           for task in data[project] :
               
               n = n + 1
               i = 1
               
               print("Estos son los valores ::" + str(n) + ", " + str(i) + "======" + project + "/////" + task)
               
               column       =sheet.cell(row=n, column=i)
               column.value = project
               i  = i + 1
               column2      =sheet.cell(row=n, column=i)
               column2.value=task
               i  = i + 1
               column3      =sheet.cell(row=n, column=i)
               column3.value=data[project][task]["MON"]
               i  = i + 1
               column4      =sheet.cell(row=n, column=i)
               column4.value=data[project][task]["TUE"]
               i  = i + 1
               column5      =sheet.cell(row=n, column=i)
               column5.value=data[project][task]["WED"]
               i  = i + 1
               column6      =sheet.cell(row=n, column=i)
               column6.value=data[project][task]["THU"]
               i  = i + 1
               column7      =sheet.cell(row=n, column=i)
               column7.value=data[project][task]["FRI"]
               i  = i + 1
               column8      =sheet.cell(row=n, column=i)
               column8.value=data[project][task]["SAT"]
               i  = i + 1
               
            
            
            
            
        
        book.save(settings.STATICFILES_USER_IMAGES_DIRS[0] + "/" + str(outputfile))
        print("######################### Llegamos a lo que esperabamos ####################")
        
    
        